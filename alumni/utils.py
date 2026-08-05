import os
from io import BytesIO
from django.conf import settings
from django.core.files.base import ContentFile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def generate_donation_receipt_pdf(donation):
    """
    Generates a professional PDF receipt for a donation using ReportLab
    and attaches it to the donation instance.
    Returns the BytesIO pdf buffer or path.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Palette
    EMERALD = colors.HexColor('#064E3B')
    GOLD = colors.HexColor('#D4AF37')
    DARK_TEXT = colors.HexColor('#1F2937')
    LIGHT_BG = colors.HexColor('#F8FAFC')
    
    # Custom Styles
    title_style = ParagraphStyle(
        'HeaderTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=EMERALD,
        alignment=1, # Center
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'HeaderSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#4B5563'),
        alignment=1,
        spaceAfter=15
    )
    receipt_tag_style = ParagraphStyle(
        'ReceiptTag',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=GOLD,
        alignment=1,
        spaceAfter=15
    )
    normal_bold = ParagraphStyle(
        'NormalBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=DARK_TEXT
    )
    normal_text = ParagraphStyle(
        'NormalText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=DARK_TEXT
    )
    footer_style = ParagraphStyle(
        'FooterText',
        parent=styles['Italic'],
        fontName='Helvetica-Oblique',
        fontSize=9,
        textColor=colors.HexColor('#6B7280'),
        alignment=1,
        spaceBefore=20
    )

    story = []

    # Institution Header
    story.append(Paragraph("ANWAR MAJMAU SHARIATH & DAWA COLLEGE ALUMNI", title_style))
    story.append(Paragraph("Vidyanagar, Nilambur, Malappuram, Kerala | Phone: +91 8606 140 996 | Email: anwarmajmau01@gmail.com", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=GOLD, spaceBefore=0, spaceAfter=15))
    
    # Receipt Title
    story.append(Paragraph("OFFICIAL DONATION RECEIPT", receipt_tag_style))
    
    # Metadata Table
    meta_data = [
        [Paragraph("<b>Receipt No:</b>", normal_bold), Paragraph(donation.receipt_number, normal_text),
         Paragraph("<b>Date:</b>", normal_bold), Paragraph(donation.date.strftime("%d %b %Y, %I:%M %p"), normal_text)],
        [Paragraph("<b>Transaction ID:</b>", normal_bold), Paragraph(donation.transaction_id or "N/A", normal_text),
         Paragraph("<b>Status:</b>", normal_bold), Paragraph(f"<font color='#064E3B'><b>{donation.receipt_status}</b></font>", normal_text)],
    ]
    meta_table = Table(meta_data, colWidths=[3.5*cm, 5*cm, 3*cm, 5*cm])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), LIGHT_BG),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TEXTCOLOR', (0,0), (-1,-1), DARK_TEXT),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#EDF2F7')),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 20))

    # Donor & Payment Details Table
    details_data = [
        [Paragraph("<b>Field</b>", normal_bold), Paragraph("<b>Details</b>", normal_bold)],
        [Paragraph("Donor Full Name", normal_bold), Paragraph(donation.donor_name, normal_text)],
        [Paragraph("Batch / Class", normal_bold), Paragraph(donation.batch, normal_text)],
        [Paragraph("Phone Number", normal_bold), Paragraph(donation.phone_number, normal_text)],
        [Paragraph("Donation Amount", normal_bold), Paragraph(f"<b>₹ {donation.amount:,.2f}</b>", ParagraphStyle('Amt', parent=normal_bold, fontSize=12, textColor=EMERALD))],
    ]
    
    details_table = Table(details_data, colWidths=[5.5*cm, 11*cm])
    details_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), EMERALD),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('LEFTPADDING', (0,0), (-1,-1), 12),
        ('RIGHTPADDING', (0,0), (-1,-1), 12),
        ('BOX', (0,0), (-1,-1), 1, EMERALD),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0,1), (-1,1), colors.white),
        ('BACKGROUND', (0,2), (-1,2), LIGHT_BG),
        ('BACKGROUND', (0,3), (-1,3), colors.white),
        ('BACKGROUND', (0,4), (-1,4), colors.HexColor('#ECFDF5')), # Subtle emerald tint for amount
    ]))
    
    # White text style for header table
    header_style = ParagraphStyle('HStyle', parent=normal_bold, textColor=colors.white)
    details_data[0] = [Paragraph("<b>Field</b>", header_style), Paragraph("<b>Details</b>", header_style)]
    
    story.append(details_table)
    story.append(Spacer(1, 30))

    # Thank you note & seal placeholder
    story.append(Paragraph("<b>Thank you for your noble contribution to Anwar Alumni Network!</b>", ParagraphStyle('ThankYou', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11, textColor=EMERALD, alignment=1)))
    story.append(Paragraph("Your contribution helps support institution development, student welfare, and alumni activities.", ParagraphStyle('NoteText', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#6B7280'), alignment=1, spaceBefore=4)))
    
    story.append(Spacer(1, 40))

    # Signature Row
    sig_data = [
        [Paragraph("_______________________<br/><b>Treasurer / Secretary</b><br/>Anwar Alumni Network", ParagraphStyle('Sig1', parent=normal_text, alignment=0)),
         Paragraph("_______________________<br/><b>Authorized Signatory</b><br/>Anwar Majmau College", ParagraphStyle('Sig2', parent=normal_text, alignment=2))]
    ]
    sig_table = Table(sig_data, colWidths=[8.25*cm, 8.25*cm])
    sig_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(sig_table)
    
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#CBD5E1'), spaceBefore=10, spaceAfter=10))
    story.append(Paragraph("This is a computer-generated receipt for Anwar Alumni Network. Copyright © 2026 Anwar Alumni Network.", footer_style))

    doc.build(story)
    pdf_value = buffer.getvalue()
    buffer.close()
    
    # Save PDF to donation instance
    filename = f"Receipt_{donation.receipt_number}.pdf"
    donation.receipt_file.save(filename, ContentFile(pdf_value), save=True)
    return pdf_value


def export_students_to_excel():
    """Export all students to Excel Workbook"""
    from alumni.models import Student
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students List"
    
    # Headers
    headers = ["Serial Number", "Full Name", "Batch", "Phone Number", "Address", "Created Date"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    students = Student.objects.all()
    for student in students:
        ws.append([
            student.serial_number,
            student.name,
            student.batch,
            student.phone_number,
            student.address,
            student.created_at.strftime("%Y-%m-%d %H:%M")
        ])
        
    # Auto adjust width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_donations_to_excel():
    """Export all donations to Excel Workbook"""
    from alumni.models import Donation
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Donations List"
    
    headers = ["Receipt No", "Date", "Donor Name", "Batch", "Phone Number", "Amount (INR)", "Transaction ID", "Status"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    donations = Donation.objects.all()
    for d in donations:
        ws.append([
            d.receipt_number,
            d.date.strftime("%Y-%m-%d %H:%M"),
            d.donor_name,
            d.batch,
            d.phone_number,
            float(d.amount),
            d.transaction_id or "N/A",
            d.receipt_status
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def import_students_from_excel(excel_file):
    """
    Imports students from uploaded Excel file.
    Expects columns: Serial Number, Name, Batch, Phone Number, Address
    Returns dict: {'created': count, 'updated': count, 'errors': list}
    """
    from alumni.models import Student
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    created_count = 0
    updated_count = 0
    errors = []
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'created': 0, 'updated': 0, 'errors': ["Uploaded file is empty"]}
        
    # Skip header if header contains 'Serial'
    first_row = [str(cell).strip().lower() for cell in rows[0] if cell is not None]
    start_index = 1 if any('serial' in c for c in first_row) else 0
    
    for row_idx, row in enumerate(rows[start_index:], start=start_index+1):
        if not row or not any(row):
            continue
            
        serial_no = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
        batch = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
        phone = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
        address = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
        
        if not serial_no or not name:
            errors.append(f"Row {row_idx}: Missing Serial Number or Name")
            continue
            
        student, created = Student.objects.update_or_create(
            serial_number=serial_no,
            defaults={
                'name': name,
                'batch': batch or "General Batch",
                'phone_number': phone or "N/A",
                'address': address or "N/A",
            }
        )
        if created:
            created_count += 1
        else:
            updated_count += 1
            
    return {
        'created': created_count,
        'updated': updated_count,
        'errors': errors
    }
